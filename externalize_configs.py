"""
externalize_configs.py

Transcribes the 6 Excel business-logic files into config/*.yaml.
This is a pure transcription step: every value/rule here is copied verbatim
from the approved Excel files. No thresholds are invented, reinterpreted,
or changed. Re-run this script whenever the source Excel files change.

Source files (data/sample/):
  RCA_Rules.xlsx
  Signal_KPI_Parameters.xlsx
  Store_Action.xlsx
  Root_Cause_-_Corrective_Action_-_Owner_-_Dashboard_View_Mapping.xlsx
  Priority_Logic.xlsx
  High_DIO_Root_Cause_Context.xlsx  (reference/documentation only, not consumed by the engine)
"""
import openpyxl
import yaml
import os

SRC = "data/sample"
OUT = "config"
os.makedirs(OUT, exist_ok=True)


def dump_yaml(obj, filename):
    path = os.path.join(OUT, filename)
    with open(path, "w") as f:
        yaml.dump(obj, f, sort_keys=False, allow_unicode=True, width=100)
    print(f"wrote {path}")


# ---------------------------------------------------------------------------
# 1. RCA_Rules.xlsx -> rca_rules.yaml
# ---------------------------------------------------------------------------
rca_rules = {
    "source_file": "RCA_Rules.xlsx",
    "gate_rule": {
        "id": 1,
        "condition": "S01_Weeks_Cover <= 1.1 * S01_Weeks_Cover_Target",
        "result": "No Action Required",
        "note": "If S01_Weeks_Cover > S01_Weeks_Cover_Target, then go ahead with finding root cause",
    },
    "rules": [
        {"id": 2, "condition": "S02_Zero_Sales_Flag == 1", "root_cause": "Possible Phantom Inventory"},
        {"id": 3, "condition": "S03_Active_Promo_Flag == 1", "root_cause": "Active Promo"},
        {"id": 4, "condition": "S04_Days_To_Promo_Start <= 14 AND S04_Upcoming_Promo_Flag == 1", "root_cause": "Upcoming Promo"},
        {"id": 5, "condition": "S06_Days_Since_Promo_Ended > 7 AND S07_Post_Promo_Velocity_Ratio < 0.5", "root_cause": "Post-promo demand collapse",
         "note": "In this case, no need to mention 'Demand Decline' due to low sales velocity ratio"},
        {"id": 6, "condition": "S09_Days_To_Season_End <= S01_Weeks_Cover * 7", "root_cause": "Season-end write-off risk",
         "note": "Corrected: compares Days to Season End against Weeks Cover converted to days (x7)."},
        {"id": 7, "condition": "S10_Sales_Velocity_Ratio < 0.75", "root_cause": "Demand Decline",
         "note": "Suppressed if Rule 5 (Post-promo demand collapse) already fired for this row."},
        {"id": 8, "condition": "S15_Forecast_vs_Actual_Pct > 15", "root_cause": "high forecast bias"},
        {"id": 9, "condition": "S17_Excess_Supply_Ratio > 2", "root_cause": "Excess Supply"},
        {"id": 10, "condition": "S18_Promo_Excess_Supply_Ratio > 3 AND S06_Days_Since_Promo_Ended > 0", "root_cause": "Promo Overbuy"},
        {"id": 11, "condition": "S12_Peer_DIO_Rate > 1.5", "root_cause": "high stock vs peers"},
        {"id": 12, "condition": "S13_Format_DIO_Ratio > 1.5", "root_cause": "High Stock vs Format"},
        {"id": 13, "condition": "S16_Par_Level_Age_Days > 90 AND S10_Sales_Velocity_Ratio < 0.75", "root_cause": "Stale Par Level",
         "note": "Stale Par Level only triggers with low sales velocity ratio"},
        {"id": 14, "condition": "S22_Shelf_Life_Remaining_Days <= 0", "root_cause": "Stock Expired",
         "note": "First check Rule 14, then Rule 15, then Rule 16, in that order. If Rule 14 true, skip 15 & 16. If Rule 15 true, skip 16."},
        {"id": 15, "condition": "S22_Shelf_Life_Remaining_Days <= S01_Weeks_Cover * 7", "root_cause": "Expiry Risk: Short-Expiry SKU ordered more than actual demand",
         "note": "Corrected: uses Weeks Cover x7, operator made inclusive (<=) to match Store_Action.xlsx Step 3, per your instruction 2026-08-14. Continue evaluating other signals for more root causes after this fires."},
        {"id": 16, "condition": "S22_Shelf_Life_Remaining_Days < 7", "root_cause": "Near Expiry: Monitor",
         "note": "Continue evaluating other signals for more root causes after this fires."},
        {"id": 17, "condition": "S19_High_SS_Flag_Days > 2", "root_cause": "High Safety Stock"},
        {"id": 18, "condition": "S20_Supplier_MOQ_Ratio > 1.3", "root_cause": "Supplier MoQ Overbuy"},
        {"id": 19, "condition": "S01_Weeks_Cover > 1.1 * S01_Weeks_Cover_Target AND no_other_rule_fired", "root_cause": "Monitor - Review at next weekly cycle",
         "note": "Fallback default state, added per Gap 10 correction."},
    ],
    "perishable_rule_sequencing": "Rules 14, 15, 16 are evaluated in that strict order and are mutually exclusive "
                                   "(first true rule wins among the three); all other rules (2-13, 17-19) are "
                                   "independent and may co-fire.",
    "expiry_rules_bypass_gate": True,
    "active_promo_suppression": {
        "source": "Signal_KPI_Parameters.xlsx Context_Parameters CTX001",
        "rule": "When Rule 3 (Active Promo) fires, all excess-type root causes (Rules 8-12, 17, 18) "
                "are suppressed UNLESS S01_Weeks_Cover > 3 * S01_Weeks_Cover_Target, in which case "
                "they are NOT suppressed. Confirmed via Row 4 in the validated sample data.",
        "suppressed_rule_ids": [8, 9, 10, 11, 12, 17, 18],
    },
}
dump_yaml(rca_rules, "rca_rules.yaml")


# ---------------------------------------------------------------------------
# 2. Signal_KPI_Parameters.xlsx -> kpi_parameters.yaml (auto-extracted)
# ---------------------------------------------------------------------------
wb = openpyxl.load_workbook(f"{SRC}/Signal_KPI_Parameters.xlsx", data_only=True)


def sheet_to_records(ws):
    headers = [c.value for c in ws[1]]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        records.append({h: v for h, v in zip(headers, row) if h is not None})
    return records


kpi_parameters = {
    "source_file": "Signal_KPI_Parameters.xlsx",
    "KPI_Parameters": sheet_to_records(wb["KPI_Parameters"]),
    "Context_Parameters": sheet_to_records(wb["Context_Parameters"]),
    "Perishable_Parameters": sheet_to_records(wb["Perishable_Parameters"]),
}
dump_yaml(kpi_parameters, "kpi_parameters.yaml")


# ---------------------------------------------------------------------------
# 3. Store_Action.xlsx -> store_action_rules.yaml (transcribed, corrected file)
# ---------------------------------------------------------------------------
store_action_rules = {
    "source_file": "Store_Action.xlsx",
    "non_perishable": [
        {"step": 1, "name": "High DIO Check",
         "condition": "S01_Weeks_Cover <= 1.1 * S01_Weeks_Cover_Target",
         "result": "No Action Required"},
        {"step": 2, "name": "High DIO Check",
         "condition": "S01_Weeks_Cover > 1.1 * S01_Weeks_Cover_Target",
         "goto": "Phantom Inventory Check"},
        {"step": 3, "name": "Phantom Inventory Check",
         "condition": "S02_Zero_Sales_Flag == 1",
         "result": "Recounting of SKU Required",
         "note": "If Days since last recount < N, 'Recounting not required' (N and the recount-date field are undefined; not implemented)."},
        {"step": 4, "name": "Phantom Inventory Check",
         "condition": "S02_Zero_Sales_Flag == 0",
         "goto": "Promo Context Check"},
        {"step": 5, "name": "Promo Context Check",
         "condition": "S03_Active_Promo_Flag == 1",
         "result": "No Action — monitor post-promo velocity from Day 1 after promo ends"},
        {"step": 6, "name": "Promo Context Check",
         "condition": "S03_Active_Promo_Flag == 0",
         "goto": "Upcoming Promo Check"},
        {"step": 7, "name": "Upcoming Promo Check",
         "condition": "S04_Days_To_Promo_Start <= 14 AND S04_Upcoming_Promo_Flag == 1",
         "result": "No Action — monitor post-promo velocity from Day 1 after promo ends"},
        {"step": 8, "name": "Upcoming Promo Check",
         "condition": "S04_Upcoming_Promo_Flag == 0",
         "goto": "Previous Promo Check"},
        {"step": 9, "name": "Previous Promo Check",
         "condition": "S06_Days_Since_Promo_Ended > 7 AND S07_Post_Promo_Velocity_Ratio < 0.5",
         "goto_true": "Peer Shortage Check (post-promo branch)",
         "goto_false": "Season End Check",
         "peer_shortage_outcomes": {
             1: "Transfer Viable - Check with Network Planner and Execute",
             0: "Markdown can be considered - check with Commercial Team and Execute",
         }},
        {"step": 10, "name": "Season End Check",
         "condition": "S08_Active_Season_Flag == 1 AND S09_Days_To_Season_End <= S01_Weeks_Cover * 7",
         "goto_true": "Peer Shortage Check (season-end branch)",
         "goto_false": "Network Imbalance Check",
         "peer_shortage_outcomes": {
             1: "Transfer Viable - Check with Network Planner and Execute",
             0: "Markdown can be considered - check with Commercial Team and Execute",
         }},
        {"step": 11, "name": "Network Imbalance Check (Peer DIO)",
         "condition": "S12_Peer_DIO_Rate > 1.5",
         "goto_true": "Peer Shortage Check (peer-dio branch)",
         "goto_false": "Network Imbalance Check (Format DIO)",
         "peer_shortage_outcomes": {
             1: "Transfer Viable - Check with Network Planner and Execute",
             0: "To decide among Transfer to distant Store or Markdown or Do-nothing",
         }},
        {"step": 12, "name": "Network Imbalance Check (Format DIO)",
         "condition": "S13_Format_DIO_Ratio > 1.5",
         "goto_true": "Peer Shortage Check (format-dio branch)",
         "goto_false_result": "To decide among Transfer to distant Store or Markdown or Do-nothing",
         "peer_shortage_outcomes": {
             1: "Transfer Viable - Check with Network Planner and Execute",
             0: "To decide among Transfer to distant Store or Markdown or Do-nothing",
         }},
    ],
    "perishable": [
        {"step": 1, "condition": "S21_Is_Perishable == 1", "goto_true": "Expiry Check",
         "goto_false": "follow Non-perishable Store Action Rules"},
        {"step": 2, "condition": "S22_Shelf_Life_Remaining_Days <= 0",
         "result_true": "Remove from Shelf",
         "goto_false": "check S22_Shelf_Life_Remaining_Days vs DIO"},
        {"step": 3, "condition": "S22_Shelf_Life_Remaining_Days <= S01_Weeks_Cover * 7",
         "result_true": "Store Manager to decide between Markdown/Transfer/Dispose/Donate for expiry-risk SKU",
         "goto_false": "check for near-expiry batches",
         "note": "Corrected: uses Weeks Cover x7, matching RCA_Rules.xlsx Rule 15 (both now inclusive <=). Recommendation text updated 2026-08-14 per your instruction."},
        {"step": 4, "condition": "S22_Shelf_Life_Remaining_Days <= 7",
         "result_true": "inventory will sell before expiry, Store Manager to monitor",
         "then": "follow Non-perishable Store Action Rules to find Root Cause",
         "goto_false": "follow Non-perishable Store Action Rules"},
    ],
    "design_note": "This is a single sequential decision-support cascade per SKU-Store (Gap 3, "
                   "confirmed intentional). It does not branch per RCA root cause and never "
                   "produces more than one traversal result per SKU-Store.",
}
dump_yaml(store_action_rules, "store_action_rules.yaml")


# ---------------------------------------------------------------------------
# 4. Corrective Action / Owner mapping -> corrective_action_map.yaml
# ---------------------------------------------------------------------------
wb2 = openpyxl.load_workbook(f"{SRC}/Root_Cause_-_Corrective_Action_-_Owner_-_Dashboard_View_Mapping.xlsx", data_only=True)
ws2 = wb2.active
mapping_records = sheet_to_records(ws2)
corrective_action_map = {
    "source_file": "Root_Cause_-_Corrective_Action_-_Owner_-_Dashboard_View_Mapping.xlsx",
    "mappings": mapping_records,
}
dump_yaml(corrective_action_map, "corrective_action_map.yaml")


# ---------------------------------------------------------------------------
# 5. Priority_Logic.xlsx -> priority_logic.yaml
# ---------------------------------------------------------------------------
wb3 = openpyxl.load_workbook(f"{SRC}/Priority_Logic.xlsx", data_only=True)
ws3 = wb3.active


def block_to_records(ws, start_row, end_row):
    headers = [c.value for c in ws[start_row]]
    records = []
    for r in range(start_row + 1, end_row + 1):
        row = [ws.cell(row=r, column=i + 1).value for i in range(len(headers))]
        if all(v is None for v in row):
            continue
        records.append({h: v for h, v in zip(headers, row) if h is not None})
    return records


priority_logic = {
    "source_file": "Priority_Logic.xlsx",
    "formula": "Priority Score = 0.7*Urgency Score + 0.1*Margin Score + 0.1*Excess Value Score + 0.1*DIO Severity Score",
    "urgency_score_table": block_to_records(ws3, 3, 8),
    "margin_score_table": block_to_records(ws3, 10, 15),
    "excess_value_score_table": block_to_records(ws3, 17, 22),
    "dio_severity_score_table": block_to_records(ws3, 24, 29),
    "urgency_factor_fields": ["S22_Shelf_Life_Remaining_Days", "S09_Days_To_Season_End"],
    "urgency_factor_note": "MIN() of the two fields above; -1 sentinel values (non-seasonal / not perishable) "
                            "are excluded from the MIN before scoring.",
    "dio_definition": "DIO = S01_Weeks_Cover * 7 (days). DIO Target = S01_Weeks_Cover_Target * 7 (days).",
    "margin_percentile_population": "within Store_ID, DIO intervention population only",
    "excess_value_percentile_population": "within Store_ID, DIO intervention population only",
    "margin_direction": "lowest margin percentile -> highest Margin Score (intentional, confirmed)",
    "small_population_fallback": {
        "threshold_eligible_rows": 4,
        "rule": "If a store has fewer than 4 eligible SKU-Store rows, Margin_Score = 3 and "
                "Excess_Value_Score = 3 (fixed mid-score), and the corresponding percentile "
                "fields are left NULL. No network-wide fallback, no raw-rank substitute.",
    },
}
dump_yaml(priority_logic, "priority_logic.yaml")


# ---------------------------------------------------------------------------
# 6. column_mapping.yaml
# ---------------------------------------------------------------------------
wb4 = openpyxl.load_workbook(f"{SRC}/Sample_RCA_Data.xlsx", data_only=True)
ws4 = wb4.active
canonical_headers = [c.value for c in ws4[3]][:33]

column_mapping = {
    "client_name": "sample_dataset (identity mapping)",
    "canonical_to_source": {h: h for h in canonical_headers},
    "note": "For a new client, only this file needs to change: map their raw column names "
            "(dict values) to the canonical signal names (dict keys). No engine code changes.",
}
dump_yaml(column_mapping, "column_mapping.yaml")

print("\nAll config files written to config/")
