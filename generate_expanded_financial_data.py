"""
generate_expanded_financial_data.py

Regenerates Financial_Impact_Data.xlsx to match the expanded
Sample_RCA_Data.xlsx (20,084 rows), using the EXACT SAME approved schema
and formulas as the original 70-row version — nothing new invented:

  SKU_ID, Store_ID                    -- join key (matches Sample_RCA_Data)
  Unit_Cost, Unit_Price                -- synthetic inputs
  Gross_Margin_Per_Unit = Price-Cost    -- formula
  Gross_Margin_Pct = GM/Price            -- formula
  Current_Stock_Units                    -- synthetic input, tied to Weeks
                                            Cover (same weekly-sales-rate
                                            methodology as before)
  Target_Stock_Units = Current * (WCT/WC) -- derived from Sample_RCA_Data's
                                             Weeks Cover ratio, same as before
  Excess_Units = MAX(Current-Target, 0)    -- formula
  Excess_Value = Excess_Units * Unit_Cost  -- formula
"""
import random
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment

random.seed(42)

df = pd.read_excel("data/sample/Sample_RCA_Data.xlsx", sheet_name=0, header=2, engine="calamine")
df = df.dropna(how="all")

cat_cost_range = {
    "Ambient": (0.60, 2.50), "BWS": (5.00, 14.00), "Bakery": (0.50, 2.00),
    "Chilled Dairy": (0.70, 2.80), "Fresh": (0.50, 2.50), "Frozen": (1.00, 4.50),
    "Health & Beauty": (1.80, 7.00), "Seasonal/Gifting": (3.00, 15.00),
}
cat_margin_range = {
    "Ambient": (0.28, 0.35), "BWS": (0.32, 0.42), "Bakery": (0.22, 0.32),
    "Chilled Dairy": (0.25, 0.33), "Fresh": (0.20, 0.30), "Frozen": (0.28, 0.36),
    "Health & Beauty": (0.38, 0.48), "Seasonal/Gifting": (0.35, 0.45),
}
cat_weekly_sales = {
    "Ambient": (40, 80), "BWS": (15, 35), "Bakery": (30, 70),
    "Chilled Dairy": (40, 90), "Fresh": (25, 60), "Frozen": (20, 45),
    "Health & Beauty": (10, 25), "Seasonal/Gifting": (5, 20),
}
format_sales_multiplier = {"Hypermarket": 2.5, "Superstore": 2.0, "Metro": 1.0, "Express": 0.6}
format_price_multiplier = {"Hypermarket": 1.00, "Superstore": 1.00, "Metro": 1.02, "Express": 1.05}

sku_cost_cache = {}
special_handling_count = 0
records = []

for row in df.to_dict(orient="records"):
    sku, store, cat, fmt = row["SKU_ID"], row["Store_ID"], row["Category"], row["Store_Format"]
    wc = float(row["S01_Weeks_Cover"])
    wct = float(row["S01_Weeks_Cover_Target"])

    cost_lo, cost_hi = cat_cost_range[cat]
    if sku not in sku_cost_cache:
        sku_cost_cache[sku] = round(random.uniform(cost_lo, cost_hi), 2)
    unit_cost = sku_cost_cache[sku]

    margin_lo, margin_hi = cat_margin_range[cat]
    margin_pct = random.uniform(margin_lo, margin_hi)
    unit_price = round((unit_cost / (1 - margin_pct)) * format_price_multiplier[fmt], 2)

    ws_lo, ws_hi = cat_weekly_sales[cat]
    weekly_sales = random.uniform(ws_lo, ws_hi) * format_sales_multiplier[fmt] * random.uniform(0.85, 1.15)

    if wc == 0:
        current_stock_units = 0
        target_stock_units = 0
        special_handling_count += 1
    else:
        current_stock_units = max(1, round(weekly_sales * wc))
        target_stock_units = round(current_stock_units * (wct / wc), 2)

    records.append((sku, store, unit_cost, unit_price, current_stock_units, target_stock_units))

print(f"Prepared {len(records)} financial records ({special_handling_count} zero-Weeks-Cover special cases)")

# ---------------- Build workbook ----------------
wb = Workbook()
ws = wb.active
ws.title = "Financial_Impact_Data"

headers = ["SKU_ID", "Store_ID", "Unit_Cost", "Unit_Price", "Gross_Margin_Per_Unit",
           "Gross_Margin_Pct", "Current_Stock_Units", "Target_Stock_Units", "Excess_Units", "Excess_Value"]

FONT_NAME = "Arial"
header_font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
input_font = Font(name=FONT_NAME, color="0000FF")
formula_font = Font(name=FONT_NAME, color="000000")
derived_static_font = Font(name=FONT_NAME, color="000000", italic=True)
plain_font = Font(name=FONT_NAME)

for col, h in enumerate(headers, start=1):
    c = ws.cell(row=1, column=col, value=h)
    c.font = header_font
    c.fill = header_fill
    c.alignment = Alignment(horizontal="center")

ws.freeze_panes = "A2"

# Bulk append (fast) instead of per-cell .cell() writes, then apply styles
# in a second, lightweight pass. Formulas reference the row they land on.
for i, (sku, store, cost, price, cur_stock, tgt_stock) in enumerate(records):
    r = i + 2
    ws.append([sku, store, cost, price, f"=D{r}-C{r}", f"=E{r}/D{r}",
               cur_stock, round(tgt_stock, 2), f"=MAX(G{r}-H{r},0)", f"=I{r}*C{r}"])

for r in range(2, len(records) + 2):
    ws.cell(row=r, column=1).font = plain_font
    ws.cell(row=r, column=2).font = plain_font
    ws.cell(row=r, column=3).font = input_font
    ws.cell(row=r, column=3).number_format = '£#,##0.00'
    ws.cell(row=r, column=4).font = input_font
    ws.cell(row=r, column=4).number_format = '£#,##0.00'
    ws.cell(row=r, column=5).font = formula_font
    ws.cell(row=r, column=5).number_format = '£#,##0.00'
    ws.cell(row=r, column=6).font = formula_font
    ws.cell(row=r, column=6).number_format = '0.0%'
    ws.cell(row=r, column=7).font = input_font
    ws.cell(row=r, column=7).number_format = '#,##0'
    ws.cell(row=r, column=8).font = derived_static_font
    ws.cell(row=r, column=8).number_format = '#,##0.0'
    ws.cell(row=r, column=9).font = formula_font
    ws.cell(row=r, column=9).number_format = '#,##0.0'
    ws.cell(row=r, column=10).font = formula_font
    ws.cell(row=r, column=10).number_format = '£#,##0.00'

widths = [16, 12, 12, 12, 20, 16, 20, 20, 14, 16]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.cell(row=1, column=8).comment = Comment(
    "Target_Stock_Units = Current_Stock_Units x (S01_Weeks_Cover_Target / S01_Weeks_Cover), "
    "calculated at generation time from Sample_RCA_Data.xlsx. Regenerate this file if "
    "Sample_RCA_Data.xlsx changes.",
    "Engine Notes", height=100, width=340,
)
ws.cell(row=1, column=7).comment = Comment(
    "Synthetic input, tied to each row's Weeks Cover so it stays internally consistent with "
    "the Weeks Cover definition (Weeks Cover = Stock / Average Weekly Sales).",
    "Engine Notes", height=90, width=340,
)

wb.save("data/sample/Financial_Impact_Data.xlsx")
print("Saved data/sample/Financial_Impact_Data.xlsx")
