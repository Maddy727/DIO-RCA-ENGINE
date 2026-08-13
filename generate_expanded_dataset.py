"""
generate_expanded_dataset.py

Generates the approved ~20,000-row synthetic operational dataset:
  6 regions, ~55 stores (assortment breadth varying by Store_Format),
  8 categories (unchanged), ~900 unique SKUs stocked at a realistic
  subset of stores, ~40% healthy (DIO <= 1.1x Target, no RCA/action)
  / ~60% breaching with varied severity.

This is DUMMY/DEMO data — unlike the original 70-row Sample_RCA_Data.xlsx,
it has NO manually-attributed ground-truth columns (AH-AL), because it is
not used to validate the engine (the engine is already validated against
the original 70-row file, which is untouched and stays as the regression
baseline). This file's only purpose is realistic demo volume/variety; the
already-validated engine computes whatever RCA/Store Action/Priority
results follow from these signal values — nothing is hand-authored here.

Canonical schema matches Sample_RCA_Data.xlsx exactly (same 33 columns),
so no engine/config changes are needed to consume it.
"""
import random
import openpyxl
from openpyxl.styles import Font, PatternFill

random.seed(42)

CATEGORIES = ["Ambient", "BWS", "Bakery", "Chilled Dairy", "Fresh", "Frozen",
              "Health & Beauty", "Seasonal/Gifting"]
PERISHABLE_CATEGORIES = {"Fresh", "Bakery", "Chilled Dairy"}
REGIONS = ["London", "South East", "South West", "Midlands", "North West", "Scotland"]

# (format, store_count, avg_skus_per_category) — drives ~20,000 total rows
FORMAT_PROFILE = [
    ("Express", 20, 24),
    ("Metro", 15, 43),
    ("Superstore", 12, 58),
    ("Hypermarket", 8, 87),
]

CATEGORY_WC_TARGET_RANGE = {
    "Ambient": (2.5, 4.0), "BWS": (3.0, 4.5), "Bakery": (1.0, 1.8),
    "Chilled Dairy": (1.2, 2.2), "Fresh": (1.0, 1.8), "Frozen": (2.5, 4.0),
    "Health & Beauty": (2.5, 4.0), "Seasonal/Gifting": (2.5, 4.5),
}

STORE_NAME_TEMPLATES = {
    "Express": ["Express {area}", "Express {area} High St", "Express {area} Local"],
    "Metro": ["Metro {area}", "Metro {area} Central"],
    "Superstore": ["Superstore {area}", "Superstore {area} Retail Park"],
    "Hypermarket": ["Hypermarket {area}", "Hypermarket {area} Extra"],
}
AREA_NAMES_BY_REGION = {
    "London": ["Angel", "Oval", "Croydon", "Ealing", "Greenwich", "Hackney", "Wembley", "Barking", "Bromley", "Sutton"],
    "South East": ["Reading", "Brighton", "Oxford", "Southampton", "Guildford", "Canterbury", "Maidstone", "Slough"],
    "South West": ["Bristol", "Exeter", "Plymouth", "Bath", "Bournemouth", "Gloucester", "Swindon", "Taunton"],
    "Midlands": ["Birmingham", "Nottingham", "Leicester", "Coventry", "Derby", "Wolverhampton", "Stoke", "Northampton"],
    "North West": ["Manchester", "Liverpool", "Preston", "Bolton", "Warrington", "Stockport", "Blackpool", "Chester"],
    "Scotland": ["Glasgow", "Edinburgh", "Aberdeen", "Dundee", "Stirling", "Inverness", "Paisley", "Perth"],
}

SKU_ADJECTIVES = ["Classic", "Fresh", "Value", "Premium", "Original", "Family", "Everyday", "Finest"]
SKU_NOUNS_BY_CATEGORY = {
    "Ambient": ["Baked Beans 4pk", "Rice Basmati 1kg", "Pasta Fusilli 500g", "Tomato Soup 400g", "Cereal Bran 750g",
                "Peanut Butter 340g", "Olive Oil 500ml", "Tinned Tuna 4pk", "Sugar White 1kg", "Coffee Instant 200g"],
    "BWS": ["Lager 18pk", "Prosecco 75cl", "Red Wine 75cl", "Cider 4pk", "Gin 70cl", "Vodka 70cl",
            "White Wine 75cl", "IPA 4pk", "Whisky 70cl", "Rose Wine 75cl"],
    "Bakery": ["White Sliced 800g", "Pain au Chocolat 4pk", "Bagels 5pk", "Wholemeal Loaf 800g",
               "Croissants 6pk", "Baguette Twin", "Bread Rolls 8pk", "Fruit Loaf 400g"],
    "Chilled Dairy": ["Whole Milk 4pt", "Cheddar Mature 400g", "Butter Unsalted 250g", "Greek Yoghurt 500g",
                       "Cream Cheese 200g", "Semi Skimmed Milk 4pt", "Mozzarella 250g", "Natural Yoghurt 1kg"],
    "Fresh": ["Mixed Veg 1kg", "Avocado Each", "Cherry Tomatoes 250g", "Bananas 6pk", "Salad Bag 200g",
              "Strawberries 400g", "Peppers 3pk", "Broccoli Each"],
    "Frozen": ["Oven Pizza Marg", "Frozen Peas 1kg", "Fish Fingers 20pk", "Ice Cream 1L", "Chips Oven 1kg",
               "Frozen Berries 500g", "Garlic Bread 2pk", "Chicken Nuggets 500g"],
    "Health & Beauty": ["Shampoo 500ml", "Toothpaste 75ml", "Hand Cream 75ml", "Body Wash 250ml",
                         "Deodorant 150ml", "Conditioner 500ml", "Face Wash 150ml", "Sun Cream 200ml"],
    "Seasonal/Gifting": ["Mother's Day Flowers", "Christmas Cracker 12pk", "Easter Egg Large", "Gift Bag Set",
                          "Advent Calendar", "Halloween Sweets 500g", "Birthday Card Pack", "Wrapping Paper 3m"],
}


def gen_sku_pool():
    """~900 unique SKUs across 8 categories, ~112 per category."""
    pool = {cat: [] for cat in CATEGORIES}
    sku_counter = 1
    for cat in CATEGORIES:
        nouns = SKU_NOUNS_BY_CATEGORY[cat]
        for i in range(112):
            adj = random.choice(SKU_ADJECTIVES)
            noun = random.choice(nouns)
            sku_id = f"SKU_{cat[:3].upper().replace(' ', '')}_{i+1:04d}"
            name = f"{adj} {noun}" if random.random() < 0.4 else noun
            pool[cat].append((sku_id, name))
            sku_counter += 1
    return pool


def gen_stores():
    stores = []
    store_counter = 1
    used_names = set()
    for fmt, count, avg_skus_per_cat in FORMAT_PROFILE:
        for _ in range(count):
            region = random.choice(REGIONS)
            area = random.choice(AREA_NAMES_BY_REGION[region])
            template = random.choice(STORE_NAME_TEMPLATES[fmt])
            store_name = template.format(area=area)
            # Guarantee uniqueness: two different stores must never share a
            # display name, since Store_Name is used as the drill-down
            # grouping key — a collision would silently merge two distinct
            # stores together in every rollup/chart.
            if store_name in used_names:
                suffix = 2
                while f"{store_name} ({suffix})" in used_names:
                    suffix += 1
                store_name = f"{store_name} ({suffix})"
            used_names.add(store_name)
            store_id = f"TS{store_counter:04d}"
            stores.append({
                "Store_ID": store_id, "Store_Name": store_name, "Store_Format": fmt,
                "Region": region, "avg_skus_per_cat": avg_skus_per_cat,
            })
            store_counter += 1
    return stores


def gen_row(row_id, sku_id, sku_name, store, category, healthy: bool):
    perishable = 1 if category in PERISHABLE_CATEGORIES else (1 if random.random() < 0.05 else 0)
    wct_lo, wct_hi = CATEGORY_WC_TARGET_RANGE[category]
    wct = round(random.uniform(wct_lo, wct_hi), 1)

    if healthy:
        wc = round(wct * random.uniform(0.5, 1.1), 2)
    else:
        severity = random.random()
        if severity < 0.5:
            wc = round(wct * random.uniform(1.15, 2.0), 2)
        elif severity < 0.85:
            wc = round(wct * random.uniform(2.0, 3.5), 2)
        else:
            wc = round(wct * random.uniform(3.5, 6.0), 2)

    zero_sales = 1 if (not healthy and random.random() < 0.08) else 0
    active_promo = 1 if random.random() < 0.10 else 0
    if active_promo:
        upcoming_promo, days_to_promo_start = 0, -1
    else:
        upcoming_promo = 1 if random.random() < 0.08 else 0
        days_to_promo_start = random.randint(1, 14) if upcoming_promo else -1
    promo_stock_ordered = random.choice([0, 1]) if (active_promo or upcoming_promo) else 0

    post_promo_collapse = (not active_promo and not healthy and random.random() < 0.10)
    if post_promo_collapse:
        days_since_promo_ended = random.randint(8, 30)
        post_promo_velocity = round(random.uniform(0.1, 0.49), 2)
    else:
        days_since_promo_ended = random.randint(-1, 7)
        post_promo_velocity = round(random.uniform(0.5, 1.3), 2)

    is_seasonal = category == "Seasonal/Gifting" and random.random() < 0.6
    active_season_flag = 1 if is_seasonal else -1
    days_to_season_end = random.randint(3, 40) if is_seasonal else -1

    sales_velocity = round(random.uniform(0.3, 0.74), 2) if (not healthy and random.random() < 0.35) else round(random.uniform(0.75, 1.6), 2)

    peer_active_promo = random.choice([0, 1])
    peer_dio_rate = round(random.uniform(0.6, 2.2), 2)
    format_dio_ratio = round(random.uniform(0.6, 2.2), 2)
    peer_shortage_flag = random.choice([0, 1])

    forecast_bias = round(random.uniform(0, 45), 1) if (not healthy and random.random() < 0.25) else round(random.uniform(0, 15), 1)
    par_level_age = round(random.uniform(0, 200), 0)
    excess_supply_ratio = round(random.uniform(2.1, 4.5), 2) if (not healthy and random.random() < 0.30) else round(random.uniform(0.5, 2.0), 2)
    promo_excess_supply_ratio = round(random.uniform(3.1, 5.0), 2) if (active_promo and random.random() < 0.3) else round(random.uniform(0.5, 3.0), 2)
    high_ss_flag_days = round(random.uniform(2.1, 6.0), 2) if (not healthy and random.random() < 0.15) else round(random.uniform(0, 2.0), 2)
    supplier_moq_ratio = round(random.uniform(1.31, 2.0), 2) if (not healthy and random.random() < 0.15) else round(random.uniform(0.8, 1.3), 2)

    if perishable:
        if not healthy and random.random() < 0.08:
            shelf_life = round(random.uniform(-2, 0), 1)
        elif not healthy and random.random() < 0.25:
            shelf_life = round(random.uniform(0.5, 14), 1)
        else:
            shelf_life = round(random.uniform(7, 45), 1)
    else:
        shelf_life = 999

    return [
        row_id, sku_id, sku_name, store["Store_ID"], store["Store_Name"], category,
        store["Store_Format"], store["Region"],
        wc, wct, zero_sales, active_promo, upcoming_promo, days_to_promo_start,
        promo_stock_ordered, days_since_promo_ended, post_promo_velocity,
        active_season_flag, days_to_season_end, sales_velocity, peer_active_promo,
        peer_dio_rate, format_dio_ratio, peer_shortage_flag, forecast_bias,
        par_level_age, excess_supply_ratio, promo_excess_supply_ratio,
        high_ss_flag_days, supplier_moq_ratio, perishable, shelf_life,
        random.randint(2, 48),
    ]


def main():
    sku_pool = gen_sku_pool()
    stores = gen_stores()

    headers = ["Row_ID", "SKU_ID", "SKU_Name", "Store_ID", "Store_Name", "Category", "Store_Format", "Region",
               "S01_Weeks_Cover", "S01_Weeks_Cover_Target", "S02_Zero_Sales_Flag", "S03_Active_Promo_Flag",
               "S04_Upcoming_Promo_Flag", "S04_Days_To_Promo_Start", "S05_Promo_Stock_Ordered_Flag",
               "S06_Days_Since_Promo_Ended", "S07_Post_Promo_Velocity_Ratio", "S08_Active_Season_Flag",
               "S09_Days_To_Season_End", "S10_Sales_Velocity_Ratio", "S11_Peer_Active_Promo_Flag",
               "S12_Peer_DIO_Rate", "S13_Format_DIO_Ratio", "S14_Peer_Shortage_Flag", "S15_Forecast_vs_Actual_Pct",
               "S16_Par_Level_Age_Days", "S17_Excess_Supply_Ratio", "S18_Promo_Excess_Supply_Ratio",
               "S19_High_SS_Flag_Days", "S20_Supplier_MOQ_Ratio", "S21_Is_Perishable", "S22_Shelf_Life_Remaining_Days",
               "Transfer_Lead_Time_Hours"]

    all_rows = []
    row_id = 1
    seen_keys = set()

    for store in stores:
        for category in CATEGORIES:
            n_skus = max(3, int(random.gauss(store["avg_skus_per_cat"], store["avg_skus_per_cat"] * 0.2)))
            available = sku_pool[category]
            chosen = random.sample(available, min(n_skus, len(available)))
            for sku_id, sku_name in chosen:
                key = (sku_id, store["Store_ID"])
                if key in seen_keys:
                    continue
                seen_keys.add(key)
                healthy = random.random() < 0.40
                row = gen_row(row_id, sku_id, sku_name, store, category, healthy)
                all_rows.append(row)
                row_id += 1

    print(f"Generated {len(all_rows)} SKU-Store rows across {len(stores)} stores, "
          f"{sum(len(v) for v in sku_pool.values())} unique SKUs, {len(REGIONS)} regions")

    healthy_count = sum(1 for r in all_rows if r[8] <= 1.1 * r[9])
    print(f"Healthy (DIO<=1.1x target): {healthy_count} ({healthy_count/len(all_rows)*100:.1f}%)")

    # ---- Write workbook, matching Sample_RCA_Data.xlsx's layout (title row 1, headers row 3, data from row 4) ----
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SKU-Store Combinations"
    ws.cell(row=1, column=1, value=f"TESCO RCA ENGINE v3 — {len(all_rows)} SKU-STORE COMBINATIONS (Synthetic Demo Dataset)")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill

    for i, row in enumerate(all_rows):
        for col, val in enumerate(row, start=1):
            ws.cell(row=4 + i, column=col, value=val)

    ws.freeze_panes = "A4"
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 16

    wb.save("data/sample/Sample_RCA_Data.xlsx")
    print("Saved data/sample/Sample_RCA_Data.xlsx")


if __name__ == "__main__":
    main()
