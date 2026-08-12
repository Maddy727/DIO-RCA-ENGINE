"""
test_rca_engine.py

Regression test: the RCA engine's output, run against every row of
Sample_RCA_Data.xlsx, must reproduce the manually-attributed ground truth
in column AH ("Root Cause Attribution") exactly. This must stay 70/70.

Run: python3 tests/test_rca_engine.py   (from project root)
"""
import os
import sys

import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "engine"))

from data_mapping import load_sku_store_data  # noqa: E402
from rca_engine import RcaEngine  # noqa: E402

DATA_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "sample", "Sample_RCA_Data.xlsx")


def _norm(s: str) -> set[str]:
    return {x.strip().lower() for x in s.split(",") if x.strip()}


def _load_ground_truth(path: str) -> dict[tuple[str, str], str]:
    """
    Ground truth (column AH, "Root Cause Attribution") is intentionally NOT
    part of the canonical input schema — it's a validation-only column, not
    real input data a client would supply. Read it directly here instead.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[3]]
    sku_col = headers.index("SKU_ID")
    store_col = headers.index("Store_ID")
    truth_col = headers.index("Root Cause Attribution")
    truth = {}
    for r in range(4, ws.max_row + 1):
        row = [ws.cell(row=r, column=i + 1).value for i in range(len(headers))]
        if all(v is None for v in row):
            continue
        truth[(row[sku_col], row[store_col])] = row[truth_col] or ""
    return truth


def run_regression():
    rows = load_sku_store_data(DATA_PATH)
    truth_lookup = _load_ground_truth(DATA_PATH)
    engine = RcaEngine()

    mismatches = []
    for row in rows:
        findings = engine.evaluate(row)
        computed = {f.root_cause for f in findings}
        truth_raw = truth_lookup.get((row["SKU_ID"], row["Store_ID"]), "")
        truth = _norm(truth_raw)
        computed_norm = _norm(", ".join(computed))
        if truth != computed_norm:
            mismatches.append({
                "row": row["SKU_ID"],
                "missing_from_computed": truth - computed_norm,
                "extra_in_computed": computed_norm - truth,
            })

    total = len(rows)
    passed = total - len(mismatches)
    print(f"RCA Engine regression: {passed}/{total} rows match ground truth")
    if mismatches:
        print("\nMismatches:")
        for m in mismatches:
            print(" ", m)
    return len(mismatches) == 0


if __name__ == "__main__":
    ok = run_regression()
    sys.exit(0 if ok else 1)
